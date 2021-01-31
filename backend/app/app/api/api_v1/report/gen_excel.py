from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment


def gen_template(header, file_name, data=[]):
    wb = Workbook()
    ws = wb.active
    ws.title = file_name
    if isinstance(header[0], str): gen_template_head_one_row(ws, header)
    if isinstance(header[0], tuple): gen_template_head_multi_row(ws, header)
    # 写入数据
    if data != []:
        start = calculate_header_rows(header)
        gen_template_with_data(ws, start, data)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def gen_template_head_one_row(ws, header):
    for i, c in enumerate(header): ws.cell(row=1, column=i + 1).value = c


def gen_template_head_multi_row(ws, header):
    x = 1
    y = 1
    for i, c in enumerate(header):
        if len(c) == 4:
            value, w, h, child = c
            merge_cells(ws, x, y, w, h, value)
            deal_child(ws, x, y + h, child)

        else:
            value, w, h = c
            merge_cells(ws, x, y, w, h, value)
        x = x + w


def deal_child(ws, x, y, child):
    for i, c in enumerate(child):
        if len(c) == 4:
            value, w, h, child = c
            merge_cells(ws, x, y, w, h, value)
            deal_child(ws, x, y + h, child)
        else:
            value, w, h = c
            merge_cells(ws, x, y, w, h, value)
        x = x + w


def merge_cells(ws, x, y, w, h, value):
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=y, start_column=x, end_row=y + h - 1, end_column=x + w - 1)
    ws.cell(row=y, column=x).value = value
    ws.cell(row=y, column=x).alignment = align


def calculate_header_rows(header):
    if isinstance(header[0], str): start = 2
    if isinstance(header[0], tuple):
        head_hight_list = []

        def dfs_head_hight(head, high=0):
            high = high + head[2]
            if len(head) == 3:
                head_hight_list.append(high)
            else:
                for h in head[3]:
                    dfs_head_hight(h, high)

        for head in header: dfs_head_hight(head)
        start = max(head_hight_list)
    return start


def gen_template_with_data(ws, start, data):
    for row_index, row in enumerate(data):
        for col_index, col in enumerate(row):
            ws.cell(row=start + row_index, column=col_index + 1).value = col


if __name__ == '__main__':
    # header [(col_name,width,height)]
    header = [
        ("机构号", 1, 3),
        ("机构名", 1, 3),
        ("员工号", 1, 3),
        ("员工名", 1, 3),
        ("A部门", 3, 1, (
            ("综合1", 2, 1, (("薪酬1", 1, 1), ("薪酬2", 1, 1))),
            ("合计", 1, 2)
        )
         ),
    ]
